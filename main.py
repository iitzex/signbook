import os
import pandas as pd
import glob
from openpyxl import load_workbook


def parse(fn):                 
	df = pd.read_html(fn)
	table = df[1]

	year = fn[9:13]
	month = fn[13:15]

	try:
		for i in range(2, 33):
		# for i in range(2, 3):
			line = table.iloc[i]
			date = line[0]
			# print('=====================', date)

			gen(year, month, date, line[2:18], '日班')
			gen(year, month, date, line[18:], '夜班')
	except IndexError as e:
		print(e)


def gen(year, month, date, human, typ):
	if typ == '日班':
		wb = load_workbook(filename = 'template.xlsx')
	else:
		wb = load_workbook(filename = f'{month}_{date}.xlsx')

	sheet = wb[typ]
	sheet[f'A4'] = f'                             日期:{year}年{month}月'
	sheet[f'C4'] = f'{date}日'
	sheet[f'G26'] = human.values[0]

	index = 6
	try:
		for k, i in enumerate(human[1:]):
			if '(' in i:
				ojt = i[:2]
				atc = i[3:5] 
			else:
				ojt = ""
				atc = i
			# print(f'D{k+index}', ojt)    
			# print(f'F{k+index}', atc) 	
			sheet[f'D{k+index}'] = ojt
			sheet[f'F{k+index}'] = atc

	except TypeError as e:
		# print(f'D{k+index}', '')    
		# print(f'F{k+index}', '!!')
		sheet[f'D{k+index}'] = ''
		sheet[f'F{k+index}'] = '!!'

	wb.save(filename = f'{month}_{date}.xlsx')

if __name__ == '__main__':
	for name in glob.glob('GR0R00005*'):
	  print(name)
	  parse(name)
	